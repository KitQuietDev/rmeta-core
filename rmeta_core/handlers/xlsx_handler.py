# handlers/xlsx_handler.py

"""
Excel Metadata Scrubber for rMeta

Rewrites spreadsheet content to a clean workbook.
Format: .xlsx
Non-destructive to cell data
"""
#
import logging
import os
from pathlib import Path
import asyncio
from typing import List
import openpyxl
from rmeta_core.utils.pii_scanner import scan_text_for_pii

logger = logging.getLogger(__name__)
__all__ = ["scrub", "get_additional_messages"]

SUPPORTED_EXTENSIONS = {"xlsx"}

# Indicates this handler supports PII detection
PII_DETECT = True

async def scrub(file_path: str) -> None:
    path = Path(file_path)
    ext = path.suffix.lower().lstrip(".")

    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {file_path}")
    if not os.access(file_path, os.R_OK | os.W_OK):
        raise PermissionError(f"Cannot access Excel file: {file_path}")
    if ext not in SUPPORTED_EXTENSIONS:
        raise ValueError(f"Unsupported file type: {ext}")

    # Make the nested function sync since openpyxl is sync
    def scrub_xlsx() -> None:
        original = openpyxl.load_workbook(file_path, data_only=True)
        clean = openpyxl.Workbook()
        clean.remove(clean.active)

        for sheet in original.worksheets:
            new_sheet = clean.create_sheet(title=sheet.title)
            for row in sheet.iter_rows(values_only=True):
                new_sheet.append(list(row))

        temp_path = path.with_suffix(".tmp.xlsx")
        clean.save(temp_path)
        os.replace(temp_path, path)
        logger.info(f"Excel scrubbed: {file_path}")

    # Run the sync function in a thread pool
    await asyncio.to_thread(scrub_xlsx)

    if not path.exists() or path.stat().st_size == 0:
        raise RuntimeError(f"Scrubbed Excel file missing or empty: {file_path}")

async def get_additional_messages(file_path: str) -> List[str]:
    messages: List[str] = [f"Metadata stripped from {Path(file_path).name}"]

    def extract_and_scan():
        wb = openpyxl.load_workbook(file_path, data_only=True)
        text_chunks: List[str] = []

        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                for cell in row:
                    if isinstance(cell, str):
                        text_chunks.append(cell)

        full_text = "\n".join(text_chunks)
        return scan_text_for_pii(full_text)

    try:
        pii_found = await asyncio.to_thread(extract_and_scan)
        for pii_type in pii_found:
            messages.append(f"PII detected: {pii_type.title()} found in file")
    except Exception as e:
        logger.warning(f"Could not scan Excel for PII: {e}")

    return messages