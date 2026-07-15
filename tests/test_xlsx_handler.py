import asyncio

import openpyxl
import pytest

from rmeta_core.handlers.xlsx_handler import scrub, get_additional_messages


def test_scrub_resets_properties_preserves_cells(dirty_xlsx):
    before = openpyxl.load_workbook(str(dirty_xlsx))
    assert before.properties.creator == "MetaMaker"

    asyncio.run(scrub(str(dirty_xlsx)))

    after = openpyxl.load_workbook(str(dirty_xlsx))
    assert after.properties.creator != "MetaMaker"
    rows = list(after.active.iter_rows(values_only=True))
    assert rows[0] == ("name", "note")
    assert rows[1][0] == "Jane Smith"


def test_get_additional_messages_flags_pii(dirty_xlsx):
    asyncio.run(scrub(str(dirty_xlsx)))
    messages = asyncio.run(get_additional_messages(str(dirty_xlsx)))
    assert any("PII detected" in m for m in messages)


def test_scrub_missing_file_raises():
    with pytest.raises(FileNotFoundError):
        asyncio.run(scrub("does-not-exist.xlsx"))
